import configparser
import json
import os
import time
from DrissionPage import ChromiumOptions, Chromium
from PyQt5.QtCore import QThread, pyqtSignal, Qt
from PyQt5.QtWidgets import QTableWidget, QFileDialog, QTableWidgetItem, QAbstractItemView
from openpyxl.reader.excel import load_workbook
from siui.components import SiLabel, SiTitledWidgetGroup, SiLongPressButton, SiOptionCardLinear, SiDenseHContainer, \
    SiDenseVContainer, SiOptionCardPlane, SiPushButton
from siui.components.button import SiSwitchRefactor, SiPushButtonRefactor
from siui.components.editbox import SiLineEdit
from siui.components.page import SiPage
from siui.components.spinbox.spinbox import SiIntSpinBox
from siui.core import SiGlobal, SiColor, Si
from config import qss
import config.CONFIG

PATH_CONFIG = config.CONFIG.CONFIG_PATH

config = configparser.ConfigParser()
config.read(PATH_CONFIG)

co = ChromiumOptions(read_file=True, ini_path=PATH_CONFIG)

try:
    config = config["chromium_options"]

    browser_path = config["browser_path"]
    broswer_address = config["address"]

except Exception as e:
    print(f"config.ini 配置文件读取失败: {e}")
    browser_path = "C:\Program Files (x86)\Microsoft\Edge\Application\msedge.exe"
    broswer_address = "127.0.0.1:9222"
finally:
    print(f"browser_path:{browser_path}")
    print(f"broswer_address:{broswer_address}")


def show_message(_type: int, title: str, text: str, icon: str):
    SiGlobal.siui.windows["MAIN_WINDOW"].LayerRightMessageSidebar().send(
        title=title,
        text=text,
        msg_type=_type,
        icon=SiGlobal.siui.iconpack.get(f"{icon}"),
        fold_after=5000)


class MainLoopThread(QThread):
    finished = pyqtSignal()

    def __init__(self, parent, browser, index_current_data):
        super().__init__(parent)
        # 初始化线程
        self.parent = parent
        self.browser = browser
        self.index_current_data = index_current_data

    def stop(self):
        self.terminate()

    def run(self):
        # run 线程
        self.last_tab = self.browser.latest_tab
        self.data = self.parent.read_to_json()
        try:
            start_index = self.index_current_data
            end_index = min(start_index + 49, len(self.data))
            if start_index >= len(self.data):
                self.finished.emit()
                return

            for i in range(start_index, end_index):
                xuehao = self.last_tab.ele(f"@id=txtstu{(i % 49) + 1}")
                score = self.last_tab.ele(f"@id=txtpoint{(i % 49) + 1}")

                print((i % 49) + 1)
                xuehao.input(self.parent.get_data_by_order(self.data, i)[0]['stu_id'])
                score.input(self.parent.get_data_by_order(self.data, i)[0]['score'])

            btus = self.last_tab.eles("@value=查询")
            for btu in btus:
                btu.click()

            self.parent.index_current_data = end_index  # 更新当前索引位置
        except Exception as e:
            print(f"发生错误: {e}")
            show_message(1, "错误", f"发生错误: {e}", "ic_fluent_error_circle_filled")
        finally:
            print("运行结束")


class Label(SiLabel):
    def __init__(self, parent, text):
        super().__init__(parent)
        self.setAlignment(Qt.AlignCenter)
        self.setFixedHeight(32)

        self.setText(text)
        self.adjustSize()
        self.resize(self.width() + 24, self.height())

    def reloadStyleSheet(self):
        self.setStyleSheet(f"color: {self.getColor(SiColor.TEXT_B)};")


class Autoexcal(SiPage):

    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)

        self.last_tab = None
        self.browser = None
        self.index_current_data: int = 0
        self.lenth: int = 0
        self.data: list = []
        self.setPadding(64)
        self.setScrollMaximumWidth(1000)
        self.setScrollAlignment(Qt.AlignLeft)
        self.setTitle("AUTOEXCAL")

        self.sheet = None

        # 创建控件组
        self.titled_widgets_group = SiTitledWidgetGroup(self)
        self.titled_widgets_group.setSiliconWidgetFlag(Si.EnableAnimationSignals)
        self.setup_set_widgets()
        # self.setup_rules_groups()
        self.setup_function_widgets()

        SiGlobal.siui.reloadStyleSheetRecursively(self)

        # 添加页脚的空白以增加美观性
        self.titled_widgets_group.addPlaceholder(64)
        # 设置控件组为页面对象
        self.setAttachment(self.titled_widgets_group)

    def setup_set_widgets(self):
        # 密堆积容器
        with self.titled_widgets_group as group:
            group.addTitle("设置")
            self.choose_boswer_btu = SiLongPressButton(self)
            self.choose_boswer_btu.resize(128, 32)
            self.choose_boswer_btu.setHint("长按选择文件夹")
            self.choose_boswer_btu.attachment().setText("选择文件夹")
            self.choose_boswer_btu.setEnabled(False)
            self.choose_boswer_btu.longPressed.connect(self.change_web_path)

            self.choose_boswer_sw = SiSwitchRefactor(self)
            self.choose_boswer_sw.toggled.connect(lambda checked: self.choose_boswer_btu.setEnabled(checked))

            boswer_filter = SiOptionCardLinear(self)
            boswer_filter.setTitle("浏览器所在文件夹", "启用以自定义选择浏览器")
            boswer_filter.load(SiGlobal.siui.iconpack.get("ic_fluent_folder_add_filled"))
            boswer_filter.addWidget(self.choose_boswer_sw)
            boswer_filter.addWidget(self.choose_boswer_btu)

            self.port_int_spin_box = SiIntSpinBox(self)
            self.port_int_spin_box.resize(128, 32)
            self.port_int_spin_box.setMaximum(65535)
            self.port_int_spin_box.setMinimum(1024)
            self.load_web_port()
            self.port_int_spin_box.setEnabled(False)
            self.port_int_spin_box.lineEdit().editingFinished.connect(self.change_web_port)

            choose_port_sw = SiSwitchRefactor(self)
            choose_port_sw.toggled.connect(lambda checked: self.port_int_spin_box.setEnabled(checked))

            choose_port_card = SiOptionCardLinear(self)
            choose_port_card.setTitle("端口号", "启用以自定义端口号")
            choose_port_card.load(SiGlobal.siui.iconpack.get("ic_fluent_plug_connected_filled"))
            choose_port_card.addWidget(choose_port_sw)
            choose_port_card.addWidget(self.port_int_spin_box)

            self.duplicate_filter_btu = SiSwitchRefactor(self)
            duplicate_filter_card = SiOptionCardLinear(self)
            duplicate_filter_card.setTitle("去重", "启用以数据去重")
            duplicate_filter_card.load(SiGlobal.siui.iconpack.get("ic_fluent_poll_off_filled"))
            duplicate_filter_card.addWidget(self.duplicate_filter_btu)

        group.addWidget(boswer_filter)
        group.addWidget(choose_port_card)
        group.addWidget(duplicate_filter_card)

    def setup_function_widgets(self):
        with self.titled_widgets_group as group:
            self.choose_switch = SiSwitchRefactor(self)
            self.choose_switch_flag = False
            data_stream_container = SiDenseHContainer(self)
            data_stream_container_v1 = SiDenseVContainer(self)
            data_stream_container_v2 = SiDenseVContainer(self)

            self.start_input = SiLineEdit(self)
            self.start_input.setTitleWidth(100)
            self.start_input.setTitle("姓名起始")
            self.start_input.setText("(9,5)")
            self.start_input.resize(350, 32)

            self.finish_input = SiLineEdit(self)
            self.finish_input.setTitleWidth(100)
            self.finish_input.setTitle("姓名结束")
            self.finish_input.setText("(200,5)")
            self.finish_input.resize(350, 32)

            self.start_input1 = SiLineEdit(self)
            self.start_input1.setTitleWidth(100)
            self.start_input1.setTitle("学号起始")
            self.start_input1.setText("(9,6)")
            self.start_input1.resize(350, 32)

            self.finish_input1 = SiLineEdit(self)
            self.finish_input1.setTitleWidth(100)
            self.finish_input1.setTitle("学号结束")
            self.finish_input1.setText("(200,6)")
            self.finish_input1.resize(350, 32)

            self.start_input2 = SiLineEdit(self)
            self.start_input2.setTitleWidth(100)
            self.start_input2.setTitle("分数起始")
            self.start_input2.setText("(9,7)")
            self.start_input2.resize(350, 32)

            self.finish_input2 = SiLineEdit(self)
            self.finish_input2.setTitleWidth(100)
            self.finish_input2.setTitle("分数结束")
            self.finish_input2.setText("(200,7)")
            self.finish_input2.resize(350, 32)

            data_stream_container_v1.addWidget(self.start_input)
            data_stream_container_v2.addWidget(self.finish_input)
            data_stream_container_v1.addWidget(self.start_input1)
            data_stream_container_v2.addWidget(self.finish_input1)
            data_stream_container_v1.addWidget(self.start_input2)
            data_stream_container_v2.addWidget(self.finish_input2)

            data_stream_container.addWidget(data_stream_container_v1)
            data_stream_container.addWidget(data_stream_container_v2)

            info_ = Label(self, "启用以自定义添加数据，若不启用，则使用默认设置")

            customize_the_input_box = SiOptionCardPlane(self)
            customize_the_input_box.setTitle("自定义输入框")
            customize_the_input_box.header().addWidget(self.choose_switch, "right")
            customize_the_input_box.body().addWidget(data_stream_container)
            customize_the_input_box.footer().addWidget(info_)
            customize_the_input_box.footer().setFixedHeight(40)
            customize_the_input_box.body().addPlaceholder(12)
            customize_the_input_box.adjustSize()

            customize_the_input_box.body().setEnabled(False)
            customize_the_input_box.footer().setEnabled(False)

            self.choose_switch.toggled.connect(lambda checked: customize_the_input_box.body().setEnabled(checked))
            self.choose_switch.toggled.connect(lambda checked: customize_the_input_box.footer().setEnabled(checked))

            self.choose_switch.toggled.connect(self.finish_input2.notifyInvalidInput)
            self.choose_switch.toggled.connect(self.finish_input1.notifyInvalidInput)
            self.choose_switch.toggled.connect(self.finish_input.notifyInvalidInput)
            self.choose_switch.toggled.connect(self.start_input.notifyInvalidInput)
            self.choose_switch.toggled.connect(self.start_input2.notifyInvalidInput)
            self.choose_switch.toggled.connect(self.start_input1.notifyInvalidInput)

            group.addWidget(customize_the_input_box)

        with self.titled_widgets_group as group:
            table_widget_height = 900
            table_widget_width = 500

            new_table_widget_height = 700
            new_table_widget_width = 500
            group.addTitle("表格数据")

            auto_input_widget_box = SiOptionCardPlane(self)
            auto_input_widget_box.adjustSize()
            auto_input_widget_box.setTitle("原始表格数据")
            auto_input_widget_box.body().setFixedSize(table_widget_height + 40, table_widget_width + 40)
            auto_input_widget_box.footer().setFixedHeight(40)

            self.table_widget = QTableWidget(self)
            self.table_widget.setStyleSheet(qss.TabelQss)
            self.table_widget.setFixedSize(table_widget_height, table_widget_width)

            self.clear_data_btu = SiLongPressButton(self)
            self.clear_data_btu.resize(80,32)
            self.clear_data_btu.attachment().setText("清除数据")
            self.clear_data_btu.longPressed.connect(self.delete_data_for_table_widget)


            choose_file_btu = SiPushButtonRefactor(self)
            choose_file_btu.setText("选择文件")
            choose_file_btu.clicked.connect(self.import_file_for_table_widget)

            auto_input_widget_box.header().addWidget(choose_file_btu, "right")
            auto_input_widget_box.body().addWidget(self.table_widget)
            auto_input_widget_box.footer().addWidget(Label(self, "使用表格数据时，请确保表格数据与输入框对应"))
            auto_input_widget_box.footer().addWidget(self.clear_data_btu, "right")

            new_input_widget_box = SiOptionCardPlane(self)
            new_input_widget_box.adjustSize()
            new_input_widget_box.setTitle("自定义表格数据")
            new_input_widget_box.body().setFixedSize(new_table_widget_height + 40, new_table_widget_width + 70)
            new_input_widget_box.footer().setFixedHeight(40)

            reload_the_data_btu = SiPushButtonRefactor(self)
            reload_the_data_btu.setText("加载数据")
            reload_the_data_btu.clicked.connect(self.reload_data_for_new_table_widget)

            new_input_widget_box.header().addWidget(reload_the_data_btu, "right")

            self.new_table_widget = QTableWidget(self)
            self.new_table_widget.setStyleSheet(qss.TabelQss)
            self.new_table_widget.setSelectionBehavior(QAbstractItemView.SelectRows)
            self.new_table_widget.setFixedSize(int(new_table_widget_height * 0.7), new_table_widget_width)
            # 设置第一列的宽度为100，第二行为200，第三行为80
            self.new_table_widget.setColumnWidth(0, 100)
            self.new_table_widget.setColumnWidth(1, 300)
            self.new_table_widget.setColumnWidth(2, 100)

            # 此容器左侧用于放置表格数据，右侧放置按钮
            operate_the_container_h = SiDenseHContainer(self)
            # 此容器用于放置表格数据
            vertical_container_for_tabular_data = SiDenseVContainer(self)
            vertical_container_for_tabular_data.addWidget(self.new_table_widget)
            # 此容器用于放置按钮
            btu_container_for_vertical_container = SiDenseVContainer(self)

            self.open_web_btu = SiPushButton(self)
            self.open_web_btu.attachment().setText("打开浏览器")
            self.open_web_btu.setFixedSize(128, 32)
            self.open_web_btu.clicked.connect(self.open_broswer)

            self.start_btu = SiPushButton(self)
            self.start_btu.attachment().setText("开始")
            self.start_btu.setFixedSize(128, 32)
            self.start_btu.clicked.connect(self.start_main_loop_in_thread)

            self.stop_btu = SiPushButton(self)
            self.stop_btu.attachment().setText("停止")
            self.stop_btu.setFixedSize(128, 32)
            self.stop_btu.clicked.connect(self.stop_main_loop_in_thread)

            self.delete_btu = SiPushButton(self)
            self.delete_btu.attachment().setText("删除")
            self.delete_btu.setFixedSize(210, 32)
            self.delete_btu.clicked.connect(self.del_data_for_new_table)

            self.insert_btu = SiPushButton(self)
            self.insert_btu.attachment().setText("插入")
            self.insert_btu.setFixedSize(210, 32)
            self.insert_btu.clicked.connect(self.insert_data_for_new_table)
            # insert data
            self.data1_input = SiLineEdit(self)
            self.data1_input.setTitleWidth(50)
            self.data1_input.setTitle("姓名")
            self.data1_input.setText("何平")
            self.data1_input.resize(210, 32)

            self.data2_input = SiLineEdit(self)
            self.data2_input.setTitleWidth(50)
            self.data2_input.setTitle("学号")
            self.data2_input.setText("2023303010311")
            self.data2_input.resize(210, 32)

            self.data3_input = SiLineEdit(self)
            # self.data3_input.setLabelWidth(100)
            self.data3_input.setTitle("分数")
            self.data3_input.setTitleWidth(50)
            self.data3_input.setText("3")
            self.data3_input.resize(210, 32)

            btu_container_for_vertical_container.addWidget(self.data1_input)
            btu_container_for_vertical_container.addWidget(self.data2_input)
            btu_container_for_vertical_container.addWidget(self.data3_input)
            btu_container_for_vertical_container.addWidget(self.insert_btu)
            btu_container_for_vertical_container.addWidget(self.delete_btu)

            temp_h = SiDenseHContainer(self)

            temp_h.addWidget(self.open_web_btu)
            temp_h.addWidget(self.start_btu)
            temp_h.addWidget(self.stop_btu)

            vertical_container_for_tabular_data.addWidget(temp_h)

            operate_the_container_h.addWidget(vertical_container_for_tabular_data)
            operate_the_container_h.addWidget(btu_container_for_vertical_container)

            new_input_widget_box.body().addWidget(operate_the_container_h)
            new_input_widget_box.footer().addWidget(Label(self, "使用表格数据时，请确保表格数据与输入框对应"))

            group.addWidget(auto_input_widget_box)
            group.addWidget(new_input_widget_box)

            # 调整父部件大小
            auto_input_widget_box.adjustSize()
            new_input_widget_box.adjustSize()
            group.adjustSize()
            self.adjustSize()

    @staticmethod
    def limit_for_table(func):
        """
        装饰器：判断self.sheet 是否 有效，有效的话执行func函数
        """

        def wrapper(self):
            if self.sheet:
                return func(self)
            else:
                print("无效，无法执行操作")
                show_message(4, "错误", "康康表格是否有问题？？？", "ic_fluent_error_circle_regular")
                return None

        return wrapper

    def delete_data_for_table_widget(self):
        self.table_widget.clear()
        self.table_widget.setRowCount(0)
        self.table_widget.setColumnCount(0)
        self.new_table_widget.clear()
        self.new_table_widget.setRowCount(0)
        self.new_table_widget.setColumnCount(0)
        self.index_current_data = 0
        os.remove(os.path.join(os.path.dirname(os.path.abspath(__file__)), 'data.json'))
        # 去实例化线程
        self.main_loop_thread = None

        show_message(2, "成功", "表格数据已清空", "ic_fluent_eraser_medium_filled")

    def import_file_for_table_widget(self):
        file_path = QFileDialog.getOpenFileName(self, "选择文件", "", "Excel Files (*.xlsx)")[0]
        if file_path:
            self.load_data_for_table_widget(file_path)
            show_message(2, "成功", "表格数据已导入", "ic_fluent_emoji_meme_filled")

    def load_data_for_table_widget(self, file_path):
        try:
            workbook = load_workbook(file_path)
            self.sheet = workbook.active
            rows = self.sheet.max_row
            cols = self.sheet.max_column

            self.table_widget.setRowCount(rows)
            self.table_widget.setColumnCount(cols)
            for row in range(rows):
                for col in range(cols):
                    cell_value = self.sheet.cell(row=row + 1, column=col + 1).value
                    item = QTableWidgetItem(str(cell_value))
                    self.table_widget.setItem(row, col, item)
        except Exception as e:
            print(e)

    @limit_for_table
    def reload_data_for_new_table_widget(self):
        self.new_table_widget.clear()
        try:
            if self.choose_switch.isChecked():
                self.name_start_strtuple = self.start_input.text()  # "(0,3)"
                self.name_end_strtuple = self.finish_input.text()

                self.xuehao_start_strtuple = self.start_input1.text()
                self.xuehao_end_strtuple = self.finish_input1.text()

                self.score_start_strtuple = self.start_input2.text()
                self.score_end_strtuple = self.finish_input2.text()

                self.name_start_int_row, self.name_start_int_col = self.name_start_strtuple.strip("()").split(',')
                self.name_end_int_row, self.name_end_int_col = self.name_end_strtuple.strip("()").split(',')
                self.xuehao_start_int_row, self.xuehao_start_int_col = self.xuehao_start_strtuple.strip("()").split(',')
                self.xuehao_end_int_row, self.xuehao_end_int_col = self.xuehao_end_strtuple.strip("()").split(',')
                self.score_start_int_row, self.score_start_int_col = self.score_start_strtuple.strip("()").split(',')
                self.score_end_int_row, self.score_end_int_col = self.score_end_strtuple.strip("()").split(',')

                # 在table_widget中加载上面的数据到new_table_widget
                # 清空 new_table_widget
                self.new_table_widget.clear()
                self.new_table_widget.setRowCount(0)
                self.new_table_widget.setColumnCount(0)

                # 获取 table_widget 的行数和列数
                table_widget_row_count = self.table_widget.rowCount()
                table_widget_col_count = self.table_widget.columnCount()

                # 获取起始和结束的行和列索引
                name_start_row = int(self.name_start_int_row) - 1
                name_end_row = int(self.name_end_int_row) - 1
                name_start_col = int(self.name_start_int_col) - 1
                name_end_col = int(self.name_end_int_col) - 1

                xuehao_start_row = int(self.xuehao_start_int_row) - 1
                xuehao_end_row = int(self.xuehao_end_int_row) - 1
                xuehao_start_col = int(self.xuehao_start_int_col) - 1
                xuehao_end_col = int(self.xuehao_end_int_col) - 1

                score_start_row = int(self.score_start_int_row) - 1
                score_end_row = int(self.score_end_int_row) - 1
                score_start_col = int(self.score_start_int_col) - 1
                score_end_col = int(self.score_end_int_col) - 1

                # 确保索引在有效范围内
                print(f"name_start_row:{name_start_row},name_end_row:{name_end_row}")
                print(f"name_start_col:{name_start_col},name_end_col:{name_end_col}")
                print(f"xuehao_start_row:{xuehao_start_row},xuehao_end_row:{xuehao_end_row}")
                print(f"xuehao_start_col:{xuehao_start_col},xuehao_end_col:{xuehao_end_col}")
                print(f"score_start_row:{score_start_row},score_end_row:{score_end_row}")
                print(f"score_start_col:{score_start_col},score_end_col:{score_end_col}")

                # 计算 new_table_widget 的行数和列数
                new_row_count = max(name_end_row - name_start_row + 1,
                                    xuehao_end_row - xuehao_start_row + 1,
                                    score_end_row - score_start_row + 1)
                new_col_count = 3  # 假设有三列：姓名、学号、分数

                # 设置 new_table_widget 的行数和列数
                self.new_table_widget.setRowCount(new_row_count)
                self.new_table_widget.setColumnCount(new_col_count)

                # 设置列标题
                self.new_table_widget.setHorizontalHeaderLabels(["姓名", "学号", "分数"])  # 复制数据到 new_table_widget
                for i in range(new_row_count):
                    # 复制姓名
                    if name_start_row + i <= name_end_row:
                        item = self.table_widget.item(name_start_row + i, name_start_col)
                        if item:
                            self.new_table_widget.setItem(i, 0, QTableWidgetItem(item.text()))

                    # 复制学号
                    if xuehao_start_row + i <= xuehao_end_row:
                        item = self.table_widget.item(xuehao_start_row + i, xuehao_start_col)
                        if item:
                            self.new_table_widget.setItem(i, 1, QTableWidgetItem(item.text()))

                    # 复制分数
                    if score_start_row + i <= score_end_row:
                        item = self.table_widget.item(score_start_row + i, score_start_col)
                        if item:
                            self.new_table_widget.setItem(i, 2, QTableWidgetItem(item.text()))

                show_message(1, "自定义", "数据复制成功", "ic_fluent_emoji_edit_filled")
            else:
                self.new_table_widget.clear()
                self.new_table_widget.setRowCount(self.table_widget.rowCount())
                self.new_table_widget.setColumnCount(3)
                self.new_table_widget.setHorizontalHeaderLabels(["姓名", "学号", "分数"])
                # 第5列是姓名
                for i in range(8, self.table_widget.rowCount()):
                    item = self.table_widget.item(i, 4)
                    if item:
                        self.new_table_widget.setItem(i - 8, 0, QTableWidgetItem(item.text()))
                # 第6列是学号
                for i in range(8, self.table_widget.rowCount()):
                    item = self.table_widget.item(i, 5)
                    if item:
                        self.new_table_widget.setItem(i - 8, 1, QTableWidgetItem(item.text()))
                # 第7列是分数
                for i in range(8, self.table_widget.rowCount()):
                    item = self.table_widget.item(i, 6)
                    if item:
                        self.new_table_widget.setItem(i - 8, 2, QTableWidgetItem(item.text()))

            for i in range(self.new_table_widget.rowCount() - 1, -1, -1):
                if self.new_table_widget.item(i, 0) is None and self.new_table_widget.item(i,
                                                                                           1) is None and self.new_table_widget.item(
                    i, 2) is None:
                    self.new_table_widget.removeRow(i)

            self.save_to_json()
            show_message(1, "默认数据", "数据复制成功", "ic_fluent_emoji_edit_filled")
        except Exception as e:
            show_message(3, "默认数据", f"数据复制失败{e}", "ic_fluent_emoji_edit_filled")

    def save_to_json(self):
        data_list = []
        unique_ids = set()
        to_remove = []

        # 获取表格数据
        row_count = self.new_table_widget.rowCount()
        names = [self.new_table_widget.item(i, 0).text() if self.new_table_widget.item(i, 0) else None for i in
                 range(row_count)]
        xuehaos = [self.new_table_widget.item(i, 1).text() if self.new_table_widget.item(i, 1) else None for i in
                   range(row_count)]
        scores = [self.new_table_widget.item(i, 2).text() if self.new_table_widget.item(i, 2) else None for i in
                  range(row_count)]

        # 构建数据列表
        for name, xuehao, score in zip(names, xuehaos, scores):
            if name is not None and xuehao is not None and score is not None:
                data = {
                    "unique_id": names.index(name),
                    "name": name,
                    "stu_id": xuehao,
                    "score": score
                }
                data_list.append(data)

        # 检查重复项并处理
        for data in data_list:
            if data['unique_id'] in unique_ids:
                if self.duplicate_filter_btu.isChecked():
                    to_remove.append(data)
                    self.new_table_widget.removeRow(data_list.index(data))
                    show_message(1, "提示", f"已删除重复项: {data}", "ic_fluent_search_filled")
                else:
                    pass

            else:
                unique_ids.add(data['unique_id'])

        # 统计数据中所有数据的个数
        print(f"数据中包含 {len(data_list)} 个数据")
        show_message(1, "提示", f"数据中包含 {len(data_list)} 个数据", "ic_fluent_task_list_ltr_filled")

        # 删除重复项
        for data in to_remove:
            data_list.remove(data)
            print(f"已删除: {data}")

        # 重新为数据列表中的内容生成唯一编号
        for idx, data in enumerate(data_list):
            data['unique_id'] = idx

        # 保存数据到 JSON 文件
        data_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'data.json')
        print(f"数据已保存到 {data_path}")
        with open(data_path, 'w', encoding='utf-8') as f:
            json.dump(data_list, f, ensure_ascii=False, indent=4)

        print("数据已保存")

    @limit_for_table
    def insert_data_for_new_table(self):
        """
        将数据插入到新表格中
        """

        name = self.data1_input.text()
        xuehao = self.data2_input.text()
        score = self.data3_input.text()
        # 插入到新表格中
        self.new_table_widget.insertRow(self.new_table_widget.rowCount())
        self.new_table_widget.setItem(self.new_table_widget.rowCount() - 1, 0, QTableWidgetItem(name))
        self.new_table_widget.setItem(self.new_table_widget.rowCount() - 1, 1, QTableWidgetItem(xuehao))
        self.new_table_widget.setItem(self.new_table_widget.rowCount() - 1, 2, QTableWidgetItem(score))
        show_message(1, "提示", f"已添加: {name}, {xuehao}, {score}", "ic_fluent_task_list_ltr_filled")
        self.save_to_json()

    @limit_for_table
    def del_data_for_new_table(self):
        """
        删除选中的当前行数据，可以多选删除
        """
        selected_rows = self.new_table_widget.selectionModel().selectedRows()
        if not selected_rows:
            show_message(3, "提示", "没有选中任何行", "ic_fluent_task_list_ltr_filled")
            return

        for row in sorted(selected_rows, reverse=True):
            self.new_table_widget.removeRow(row.row())
            print(f"已删除: {row.row()}")

        # 使用最后一个删除的行来显示消息
        last_deleted_row = selected_rows[-1]
        show_message(3, "提示", f"已删除: {last_deleted_row.row() + 1}", "ic_fluent_task_list_ltr_filled")
        self.save_to_json()

    @limit_for_table
    def open_broswer(self):
        try:
            if self.choose_boswer_sw.isChecked():
                self.browser = Chromium(int(self.port_int_spin_box.value()))
            else:
                self.browser = Chromium(broswer_address)
        except Exception as e:
            print(f"无法启动浏览器: {e}")
            show_message(3, "提示", f"无法启动浏览器: {e}", "ic_fluent_task_list_ltr_filled")
            return

    def read_to_json(self) -> list:
        """
        从指定的JSON文件中读取数据并返回一个列表
        :return: 包含JSON数据的列表
        """
        file_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'data.json')
        try:
            with open(file_path, 'r', encoding='utf-8') as file:
                data = json.load(file)
                return data
        except FileNotFoundError:
            print(f"文件未找到: {file_path}")
            return []
        except json.JSONDecodeError:
            print(f"JSON解码错误: {file_path}")
            return []

    def get_data_by_order(self, data, order) -> list:
        """
        根据指定的order值筛选数据
        :param data: 包含JSON数据的列表
        :param order: 要筛选的order值
        :return: 筛选出的数据列表
        """
        filtered_data = [item for item in data if item.get('unique_id') == order]
        return filtered_data

    @limit_for_table
    def main_loop(self):
        """
        主循环，从当前索引位置开始输入49个数据
        """
        if self.browser:
            self.last_tab = self.browser.latest_tab
            self.data = self.read_to_json()
            try:
                start_index = self.index_current_data
                end_index = min(start_index + 49, len(self.data))
                print(f": e1{start_index + 49},{len(self.data)}")
                if start_index >= len(self.data):
                    show_message(3, "提示", "数据已全部输入完毕", "ic_fluent_checkmark_starburst_filled")
                    return
                self.start_btu.attachment().setText(f"继续{(end_index % 49) + 1}")

                for i in range(start_index, end_index):
                    xuehao = self.last_tab.ele(f"@id=txtstu{(i % 49) + 1}")
                    score = self.last_tab.ele(f"@id=txtpoint{(i % 49) + 1}")

                    print((i % 49) + 1)
                    xuehao.input(self.get_data_by_order(self.data, i)[0]['stu_id'])
                    time.sleep(0.1)
                    score.input(self.get_data_by_order(self.data, i)[0]['score'])
                    time.sleep(0.1)
                btus = self.last_tab.eles("@value=查询")
                for btu in btus:
                    time.sleep(0.1)
                    btu.click()

                self.index_current_data = end_index  # 更新当前索引位置
            except Exception as e:
                print(f"发生错误: {e}")
                show_message(1, "错误", f"发生错误: {e}", "ic_fluent_error_circle_filled")
        else:
            show_message(3, "提示", "请先打开浏览器", "ic_fluent_error_circle_filled")

    @limit_for_table
    def start_main_loop_in_thread(self):
        # 实例化线程
        self.main_loop_thread = MainLoopThread(self, self.browser, self.index_current_data)
        self.main_loop_thread.finished.connect(self.on_main_loop_finished)
        self.main_loop_thread.start()

    @limit_for_table
    def stop_main_loop_in_thread(self):
        if self.main_loop_thread.isRunning():
            self.main_loop_thread.stop()


    def on_main_loop_finished(self):
        self.start_btu.attachment().setText("开始")
        show_message(3, "提示", "数据已全部输入完毕", "ic_fluent_checkmark_starburst_filled")
        self.start_btu.setEnabled(True)

    def change_web_path(self):
        try:
            file_path = QFileDialog.getOpenFileName(self, "选择浏览器路径", "", "Executable Files (*.exe)")[0]
            if file_path:
                config2 = configparser.ConfigParser()
                config2.read(PATH_CONFIG)
                config = config2["chromium_options"]
                config["browser_path"] = file_path
                with open(PATH_CONFIG, 'w') as configfile:
                    config2.write(configfile)
                show_message(1, "提示", "浏览器路径已更改", "ic_fluent_wrench_settings_filled")
            print(file_path)
        except Exception as e:
            print(f"发生错误: {e}")
            show_message(1, "错误", f"发生错误: {e}", "ic_fluent_error_circle_filled")

    def change_web_port(self):
        try:
            port = self.port_int_spin_box.value()
            config1 = configparser.ConfigParser()
            config1.read(PATH_CONFIG)
            config = config1["chromium_options"]
            config["address"] = f"127.0.0.1:{port}"
            with open(PATH_CONFIG, 'w') as configfile:
                config1.write(configfile)
            show_message(1, "提示", "端口已更改", "ic_fluent_wrench_settings_filled")
        except Exception as e:
            print(f"发生错误: {e}")
            show_message(1, "错误", f"发生错误: {e}", "ic_fluent_error_circle_filled")

    def load_web_port(self):
        try:
            config1 = configparser.ConfigParser()
            config1.read(PATH_CONFIG)
            config = config1["chromium_options"]
            port = config["address"].split(":")[1]
            self.port_int_spin_box.setValue(int(port))
        except Exception as e:
            print(f"发生错误: {e}")
            self.port_int_spin_box.setValue(9222)