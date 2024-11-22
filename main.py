import json
import os
import re
import time
import sys
import pyautogui

from PyQt5 import QtWidgets
from PyQt5.QtCore import QThread, pyqtSignal, Qt, QTranslator, QPoint
from PyQt5.QtGui import QMouseEvent, QColor
from PyQt5.QtWidgets import QApplication, QTableWidgetItem, QLineEdit, QAbstractItemView, QVBoxLayout, QLabel, \
    QHBoxLayout, QFileDialog, QWidget

from setting_interface import SettingInterface
from config import cfg
from new import Ui_Form
from openpyxl.reader.excel import load_workbook
from pynput import mouse

from qfluentwidgets import setThemeColor, Dialog, Flyout, InfoBarIcon, \
    FlyoutAnimationType, PrimaryPushButton, LineEdit, TransparentPushButton, ToolTipFilter, \
    ToolTipPosition, FluentTranslator, isDarkTheme, BodyLabel

from qframelesswindow import StandardTitleBar, AcrylicWindow, FramelessWindow


class detect_mouse_double_click(QThread):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.running = False
        self.paused = False
        self.double_click_count = 0
        self.double_click_threshold = 0.3  # 双击时间阈值（秒）
        self.last_click_time = 0
        self.parent = parent
        self.listener = None

    def double_click(self, x, y, button, pressed):
        if pressed and button == mouse.Button.left:
            current_time = time.time()
            if current_time - self.last_click_time < self.double_click_threshold:
                self.double_click_count += 1
                if self.double_click_count == 2:
                    self.double_click_count = 0
                    self.last_click_time = current_time
                    self.parent.on_double_click()
            else:
                self.double_click_count = 1
                self.last_click_time = current_time

    def run(self):
        self.running = True
        while self.running:
            if not self.paused:
                if self.listener is None:
                    self.listener = mouse.Listener(on_click=self.double_click)
                    self.listener.start()
                time.sleep(1)  # 使用 time.sleep 替代 join
            else:
                if self.listener is not None:
                    self.listener.stop()
                    self.listener = None
                time.sleep(0.1)

    def stop(self):
        self.running = False
        if self.listener is not None:
            self.listener.stop()


class CustomFlyoutView(QWidget):
    # 定义一个信号，用于传递输入数据
    data_submitted = pyqtSignal(str, str, str)

    def __init__(self, parent=None):
        super().__init__(parent)
        self.setFixedSize(300, 200)
        # self.setTitleBar(StandardTitleBar(self))
        # self.titleBar.raise_()
        # 创建布局管理器
        self.vBoxLayout = QVBoxLayout(self)

        # 创建姓名输入框
        self.name_label = BodyLabel('姓名:')
        self.name_input = LineEdit()
        self.name_input.setText("何平")
        self.name_input.setPlaceholderText("何平")
        self.name_box = QHBoxLayout()
        self.name_box.addWidget(self.name_label)
        self.name_box.addWidget(self.name_input)

        # 创建学号输入框
        self.id_label = BodyLabel('学号:')
        self.id_input = LineEdit()
        self.id_input.setText("2023303010311")
        self.id_input.setPlaceholderText("2023303010311")
        self.id_box = QHBoxLayout()
        self.id_box.addWidget(self.id_label)
        self.id_box.addWidget(self.id_input)

        # 创建操行分输入框
        self.score_label = BodyLabel('操行分:')
        self.score_input = LineEdit()
        self.score_input.setText("2")
        self.score_input.setPlaceholderText("2")
        self.score_box = QHBoxLayout()
        self.score_box.addWidget(self.score_label)
        self.score_box.addWidget(self.score_input)

        # 创建确认和取消按钮
        self.button_box = QHBoxLayout()
        self.ok_btu = PrimaryPushButton()
        self.ok_btu.setText("确定")
        self.ok_btu.clicked.connect(self.submit_data)
        self.cancel_btu = TransparentPushButton()
        self.cancel_btu.setText("取消")
        self.cancel_btu.clicked.connect(self.close)
        self.button_box.addWidget(self.ok_btu)
        self.button_box.addWidget(self.cancel_btu)

        # 添加控件到布局
        self.vBoxLayout.addLayout(self.name_box)
        self.vBoxLayout.addLayout(self.id_box)
        self.vBoxLayout.addLayout(self.score_box)
        self.vBoxLayout.addLayout(self.button_box)

        # 设置布局间距和边距
        self.vBoxLayout.setSpacing(12)
        self.vBoxLayout.setContentsMargins(20, 16, 20, 16)

    def submit_data(self):
        name = self.name_input.text()
        id_number = self.id_input.text()
        score = self.score_input.text()
        self.data_submitted.emit(name, id_number, score)


class SettingMyFlyout(FramelessWindow):
    def __init__(self, parent=None):
        super().__init__(parent=parent)
        self.setTitleBar(StandardTitleBar(self))
        self.titleBar.raise_()
        self.resize(800, 500)
        self.hBoxLayout = QHBoxLayout(self)
        self.settingInterface = SettingInterface(self)
        self.hBoxLayout.setContentsMargins(0, 0, 0, 0)
        self.hBoxLayout.addWidget(self.settingInterface)

        self.setQss()
        cfg.themeChanged.connect(self.setQss)
        self.show()

    def setQss(self):
        theme = 'dark' if isDarkTheme() else 'light'
        with open(f'resource/qss/{theme}/demo.qss', encoding='utf-8') as f:
            self.setStyleSheet(f.read())

    def mouseMoveEvent(self, e: QMouseEvent):  # 重写移动事件
        self._endPos = e.pos() - self._startPos
        self.move(self.pos() + self._endPos)

    def mousePressEvent(self, e: QMouseEvent):
        if e.button() == Qt.LeftButton:
            self._isTracking = True
            self._startPos = QPoint(e.x(), e.y())
        else:
            self.close()

    def mouseReleaseEvent(self, e: QMouseEvent):
        if e.button() == Qt.RightButton:
            self.close()


class MyWindow(AcrylicWindow, Ui_Form):
    def __init__(self):
        super().__init__()

        self.setupUi(self)
        self.setTitleBar(StandardTitleBar(self))
        self.titleBar.raise_()
        self.resize(1400, 900)
        # self.setWindowTitle("自定义")
        rect = QApplication.desktop().availableGeometry()
        # 居中显示窗口
        self.move(rect.width() // 2 - self.width() // 2, rect.height() // 2 - self.height() // 2)
        setThemeColor("#28afe9")
        # setTheme(Theme.DARK)
        # self.windowEffect.setMicaEffect(self.winId())
        self.initui()
        self.solt()
        # 禁用失败安全机制
        pyautogui.FAILSAFE = False

        self.tableWidget_2.verticalHeader().setVisible(False)  # 隐藏垂直表头
        self.tableWidget_2.horizontalHeader().setVisible(False)  # 隐藏水平表头
        self.tableWidget.setSelectionMode(QAbstractItemView.ExtendedSelection)
        self.tableWidget_2.setSelectionMode(QAbstractItemView.ExtendedSelection)

        self.sheet = None
        self.mouse_thread = None
        self.tool_window = None
        self.heping = 1
        self.base_table_row = 0
        self.base_table_col = 0
        self.error = 0
        self.success = 1
        self.warning = 2
        self.info = 3

    def initui(self):
        self.tool_btu = self.toolButton
        self.start_btu = self.pushButton_4
        self.import_file_btu = self.pushButton_7
        self.restart_btu = self.pushButton
        self.default_btu = self.pushButton_2
        self.remove_data_in_table2_btu = self.pushButton_6
        self.add_data_in_table2_btu = self.pushButton_5

        self.name_input = self.lineEdit
        self.name_input.setToolTip("填人名儿")
        self.name_input.installEventFilter(ToolTipFilter(self.name_input, showDelay=300, position=ToolTipPosition.TOP))
        self.name_input.setPlaceholderText("(9, 5)")
        self.name_input.setText("(9, 5)")
        self.name_input.focusInEvent = lambda envet: self.input_focus_in(self.name_input, envet)

        self.xuehao_input = self.lineEdit_2
        self.xuehao_input.setPlaceholderText("(9, 6)")
        self.xuehao_input.setText("(9, 6)")
        self.xuehao_input.focusInEvent = lambda envet: self.input_focus_in(self.xuehao_input, envet)

        self.score_input = self.lineEdit_3
        self.score_input.setPlaceholderText("(9, 7)")
        self.score_input.setText("(9, 7)")
        self.score_input.focusInEvent = lambda envet: self.input_focus_in(self.score_input, envet)

        self.groupBox.toggled.connect(self.groupBox_state_change)
        self.reload_data_btu = self.pushButton_3

    def solt(self):
        self.tool_btu.clicked.connect(self.tool_btu_click)
        self.import_file_btu.clicked.connect(self.import_file_click)
        self.restart_btu.clicked.connect(self.restart_btu_click)
        self.default_btu.clicked.connect(self.default_click)
        self.reload_data_btu.clicked.connect(self.reload_data)
        self.remove_data_in_table2_btu.clicked.connect(self.remove_data_in_table2)
        self.add_data_in_table2_btu.clicked.connect(self.add_one_data_in_table2)
        self.start_btu.clicked.connect(self.start_or_stop)
        # self.tableWidget_2.itemChanged.connect(self.save_to_json)

    def remove_data_in_table2(self):
        if self.sheet:
            # 收集所有需要删除的行号
            rows_to_remove = set()
            items = self.tableWidget_2.selectedItems()
            for item in items:
                row = item.row()
                rows_to_remove.add(row)

            # 按行号从大到小排序，避免删除行后影响其他行号
            rows_to_remove = sorted(rows_to_remove, reverse=True)

            # 统一删除所有行
            for row in rows_to_remove:
                cell_value = self.tableWidget_2.item(row, 0).text()
                self.tableWidget_2.removeRow(row)
                self.info_box("WARRING", f"已删除{cell_value}")

            self.save_to_json()
        else:
            self.info_box("WARRING", "请先导入数据")
            self.info_bar(0, "请先导入数据", "请先导入数据")

    def add_one_data_in_table2(self):
        if not self.sheet:
            self.info_bar(0, "请先导入数据", "请先导入数据")
            return
        flyout_view = CustomFlyoutView(self)
        # 连接信号和槽
        flyout_view.data_submitted.connect(self.rev_data_to_flyoutview)
        Flyout.make(flyout_view, self.add_data_in_table2_btu, self, aniType=FlyoutAnimationType.SLIDE_RIGHT)

    def rev_data_to_flyoutview(self, name, id_number, score):
        print(f"姓名: {name}, 学号: {id_number}, 操行分: {score}")
        if name and id_number and score and self.sheet:
            self.tableWidget_2.insertRow(self.tableWidget_2.rowCount())
            self.tableWidget_2.setItem(self.tableWidget_2.rowCount() - 1, 0, QTableWidgetItem(name))
            self.tableWidget_2.setItem(self.tableWidget_2.rowCount() - 1, 1, QTableWidgetItem(id_number))
            self.tableWidget_2.setItem(self.tableWidget_2.rowCount() - 1, 2, QTableWidgetItem(score))
            self.save_to_json()
            self.info_bar(1, "添加成功", "添加成功")
        else:
            self.info_bar(0, "请输入完整数据", "请输入完整数据")

    def reload_data(self):
        if self.sheet:
            matche_1 = re.findall(r'\d+', self.name_input.text())
            integer_1 = [int(match) for match in matche_1]
            self.name_start_row, self.name_start_col = integer_1
            name_tuple = (self.name_start_row, self.name_start_col)
            print(integer_1)

            matche_2 = re.findall(r'\d+', self.xuehao_input.text())
            integer_2 = [int(match) for match in matche_2]
            self.xuehao_start_row, self.xuehao_start_col = integer_2
            xuehao_tuple = (self.xuehao_start_row, self.xuehao_start_col)
            print(integer_2)

            matche_3 = re.findall(r'\d+', self.score_input.text())
            integer_3 = [int(match) for match in matche_3]
            self.score_start_row, self.score_start_col = integer_3
            score_tuple = (self.score_start_row, self.score_start_col)
            print(integer_3)

            # 感觉没啥必要写个函数，全局就三次用法

            if self.name_start_row != self.xuehao_start_row or self.name_start_row != self.score_start_row:
                self.info_bar(self.error, "错误", "请确保三个数据列的行号一致")
                return
            if self.name_start_row == 0 or self.xuehao_start_row == 0 or self.score_start_row == 0:
                self.info_bar(self.error, "错误", "请确保三个数据列的行号不为0")
                return
            self.search_data_in_table_to_table2(name_tuple, xuehao_tuple, score_tuple)
        else:
            self.info_box("错误", "请先导入数据")

    def tool_btu_click(self):
        self.tool_window = SettingMyFlyout()
        self.tool_window.raise_()

    def default_click(self):
        self.info_bar(self.info, "默认设置", "默认设置")
        self.name_input.setText("(9, 5)")
        self.xuehao_input.setText("(9, 6)")
        self.score_input.setText("(9, 7)")

    def groupBox_state_change(self):
        if self.groupBox.isChecked():
            self.tableWidget.cellClicked.connect(self.table_clicked)
        else:
            self.tableWidget.cellClicked.disconnect()

    def start_or_stop(self) -> None:
        try:
            if self.sheet:
                if self.mouse_thread is None:
                    self.mouse_thread = detect_mouse_double_click(parent=self)
                    self.mouse_thread.start()
                    self.start_btu.setText("结束")
                else:
                    self.mouse_thread.stop()
                    self.mouse_thread = None
                    self.start_btu.setText("开始")
            else:
                self.info_bar(0, "不选文件??", "先选择文件!!")
        except Exception as e:
            print("错误", e)
            self.info_bar(0, "有错误(><) !!", f"{e}")

    def on_double_click(self):
        data = self.read_to_json()
        id = self.get_data_by_order(data, self.heping)
        id_dict_value = id[0]["unique_id"]
        col_conduct_points_dict_value = id[0]["score"]
        name_dict_value = id[0]["name"]
        print(f"{id_dict_value}.{col_conduct_points_dict_value},{name_dict_value}")
        self.info_bar(1, "已输入", f"{id_dict_value}.{col_conduct_points_dict_value},{name_dict_value}")
        self.info_box(f"{id}已输入", f"{id_dict_value}.{col_conduct_points_dict_value},{name_dict_value}")

        self.heping = self.heping + 1

        # 双击时光标位置输入学号
        pyautogui.typewrite(str(id_dict_value))
        # 获取双击时光标位置
        x, y = pyautogui.position()
        # 将光标移动到左上角
        pyautogui.moveTo(0, 0)
        time.sleep(0.1)
        pyautogui.press('tab')
        time.sleep(0.1)
        pyautogui.press('enter')
        time.sleep(0.1)
        # 判断分数构成，模拟对应的数字键盘
        if len(str(col_conduct_points_dict_value)) == 1:
            pyautogui.press(str(col_conduct_points_dict_value))
        elif len(str(col_conduct_points_dict_value)) == 3:
            pyautogui.press(str(col_conduct_points_dict_value)[0])
            print(str(col_conduct_points_dict_value)[0])
            time.sleep(0.1)
            pyautogui.press(".")
            print(".")
            time.sleep(0.1)
            pyautogui.press(str(col_conduct_points_dict_value)[2])
            print(str(col_conduct_points_dict_value)[2])
        # 再次移动到双击位置
        pyautogui.moveTo(x, y + 60)

    def import_file_click(self):

        file_path_temp1, _ = QFileDialog.getOpenFileName(self, "选择文件", "/", "Excel文件 (*.xlsx *.xls)")

        if file_path_temp1:
            self.load_data_to_table(file_path_temp1)

    def restart_btu_click(self):
        self.info_box("WAR", "RESET")
        self.tableWidget_2.clear()
        self.tableWidget.clear()
        self.sheet = None
        self.mouse_thread = None
        self.heping = 1
        self.base_table_row = 0
        self.base_table_col = 0
        if os.path.exists("data.json"):
            os.remove("data.json")

    def load_data_to_table(self, file_path):
        try:
            workbook = load_workbook(file_path)
            self.sheet = workbook.active
            rows = self.sheet.max_row
            cols = self.sheet.max_column

            self.tableWidget.setRowCount(rows)
            self.tableWidget.setColumnCount(cols)
            for row in range(rows):
                for col in range(cols):
                    cell_value = self.sheet.cell(row=row + 1, column=col + 1).value
                    item = QTableWidgetItem(str(cell_value))
                    self.tableWidget.setItem(row, col, item)

            self.search_data_in_table_to_table2()

        except Exception as e:
            print(e)

    def search_data_in_table_to_table2(self, name: tuple = (9, 5), xuehao: tuple = (9, 6), score: tuple = (9, 7)):
        name_row, name_col = name
        xuehao_row, xuehao_col = xuehao
        score_row, score_col = score

        self.tableWidget_2.setRowCount(self.sheet.max_row - name_row + 1)
        self.tableWidget_2.setColumnCount(3)

        for row, i in zip(range(name_row, self.sheet.max_row + 1), range(self.sheet.max_row)):
            item = QTableWidgetItem(str(self.sheet.cell(row=row, column=name_col).value))
            print(self.sheet.cell(row=row, column=name_col).value)
            self.tableWidget_2.setItem(i, 0, item)
        for row, i in zip(range(xuehao_row, self.sheet.max_row + 1), range(self.sheet.max_row)):
            item = QTableWidgetItem(str(self.sheet.cell(row=row, column=xuehao_col).value))
            self.tableWidget_2.setItem(i, 1, item)
        for row, i in zip(range(score_row, self.sheet.max_row + 1), range(self.sheet.max_row)):
            item = QTableWidgetItem(str(self.sheet.cell(row=row, column=score_col).value))
            self.tableWidget_2.setItem(i, 2, item)
        self.info_bar(1, "已读取数据", "已读取")

        self.save_to_json()

    def table_clicked(self, row, column):
        cell_value = self.tableWidget.item(row, column).text()
        print(f"点击了单元格 ({row + 1}, {column + 1})，内容为: {cell_value}")
        self.base_table_row = row + 1
        self.base_table_col = column + 1
        self.info_box("INFO", f"当前选中的单元格为：({row + 1},{column + 1})")

    def input_focus_in(self, type, event):
        def type_focus(mode):
            mode.setText(f"{self.base_table_row, self.base_table_col}")
            QLineEdit.focusInEvent(mode, event)

        if type == self.name_input:
            type_focus(self.name_input)
        elif type == self.xuehao_input:
            type_focus(self.xuehao_input)
        elif type == self.score_input:
            type_focus(self.score_input)

    def save_to_json(self):
        data_list = []
        unique_ids = set()
        to_remove = []

        # 获取表格数据
        row_count = self.tableWidget_2.rowCount()
        names = [self.tableWidget_2.item(i, 0).text() if self.tableWidget_2.item(i, 0) else None for i in
                 range(row_count)]
        xuehaos = [self.tableWidget_2.item(i, 1).text() if self.tableWidget_2.item(i, 1) else None for i in
                   range(row_count)]
        scores = [self.tableWidget_2.item(i, 2).text() if self.tableWidget_2.item(i, 2) else None for i in
                  range(row_count)]

        # 构建数据列表
        for name, xuehao, score in zip(names, xuehaos, scores):
            if name is not None and xuehao is not None and score is not None:
                data = {
                    "unique_id": names.index(name),
                    "name": name,
                    "stu_id": xuehao,
                    "score": score
                }
                data_list.append(data)

        # 检查重复项并处理
        for data in data_list:
            if data['unique_id'] in unique_ids:
                temp_data = data.copy()

                w = Dialog("重复编号", f'检测到表中含有重复<br/> '
                                       f'姓名：{temp_data["name"]}    <br/>'
                                       f'学号：{temp_data["stu_id"]}  <br/>'
                                       f'操行：{temp_data["score"]}   分<br/>'
                                       f'是否删除? 请注意!取消的话并不会删除数据')
                if w.exec():
                    to_remove.append(data)
                    self.tableWidget_2.removeRow(data_list.index(data))
                    self.info_bar(2, "重复编号",
                                  f'姓名：{temp_data["name"]}    <br/>'
                                  f'学号：{temp_data["stu_id"]}  <br/>'
                                  f'操行：{temp_data["score"]}   分<br/>'
                                  )
                    print(f"已删除{temp_data}")
                else:
                    print("取消删除")
                    self.info_bar(2, "已删除", "已取消删除")
            else:
                unique_ids.add(data['unique_id'])

        # 删除重复项
        for data in to_remove:
            data_list.remove(data)
            print(f"已删除: {data}")

        # 重新为数据列表中的内容生成唯一编号
        for idx, data in enumerate(data_list):
            data['unique_id'] = idx

        # 保存数据到 JSON 文件
        with open('data.json', 'w') as f:
            json.dump(data_list, f, ensure_ascii=False, indent=4)

        # 显示完成提示
        self.info_box("OK------->", f"保存成功,一共{len(data_list)}个数据")
        self.info_bar(1, "🆗啦", f"保存成功,一共{len(data_list)}个数据")
        print("数据已保存")

    def read_to_json(self) -> list:
        """
        从指定的JSON文件中读取数据并返回一个列表
        :return: 包含JSON数据的列表
        """
        file_path = "data.json"
        try:
            with open(file_path, 'r') as file:
                data = json.load(file)
                return data
        except FileNotFoundError:
            print(f"文件未找到: {file_path}")
            return []
        except json.JSONDecodeError:
            print(f"JSON解码错误: {file_path}")
            return []

    def get_data_by_order(self, data, order) -> list:
        """
        根据指定的order值筛选数据

        :param data: 包含JSON数据的列表
        :param order: 要筛选的order值
        :return: 筛选出的数据列表
        """
        filtered_data = [item for item in data if item.get('unique_id') == order]
        return filtered_data

    def info_bar(self, type: int, til: str, msg: str):
        def show_info_bar(method):
            Flyout.create(
                icon=method,
                title=self.tr(til),
                content=self.tr(msg),
                target=self,
                parent=self,
                isClosable=True,
                aniType=FlyoutAnimationType.PULL_UP
            )

        if type == 0:
            show_info_bar(InfoBarIcon.ERROR)
        elif type == 1:
            show_info_bar(InfoBarIcon.SUCCESS)
        elif type == 2:
            show_info_bar(InfoBarIcon.WARNING)
        elif type == 3:
            show_info_bar(InfoBarIcon.INFORMATION)

    def info_box(self, level: str, msg: str):
        """
        打印命令行输出
        :param level:
        :param msg: 消息
        :return: None
        """
        self.textBrowser.append(f"[{self.tr(level)}]/>{self.tr(msg)}")
        self.textBrowser.ensureCursorVisible()

    def close(self):
        if self.tool_window:
            self.tool_window.close()
        super().close()


if __name__ == '__main__':
    QApplication.setHighDpiScaleFactorRoundingPolicy(Qt.HighDpiScaleFactorRoundingPolicy.PassThrough)
    QtWidgets.QApplication.setAttribute(Qt.AA_EnableHighDpiScaling)
    QtWidgets.QApplication.setAttribute(Qt.AA_UseHighDpiPixmaps)
    QApplication.setAttribute(Qt.AA_UseHighDpiPixmaps)

    if cfg.get(cfg.dpiScale) == "Auto":
        QApplication.setHighDpiScaleFactorRoundingPolicy(
            Qt.HighDpiScaleFactorRoundingPolicy.PassThrough)
        QApplication.setAttribute(Qt.AA_EnableHighDpiScaling)
    else:
        os.environ["QT_ENABLE_HIGHDPI_SCALING"] = "0"
        os.environ["QT_SCALE_FACTOR"] = str(cfg.get(cfg.dpiScale))
    QApplication.setAttribute(Qt.AA_UseHighDpiPixmaps)
    # create application
    app = QApplication(sys.argv)
    app.setAttribute(Qt.AA_DontCreateNativeWidgetSiblings)
    # internationalization
    locale = cfg.get(cfg.language).value
    fluentTranslator = FluentTranslator(locale)
    settingTranslator = QTranslator()
    settingTranslator.load(locale, "settings", ".", "resource/i18n")
    app.installTranslator(fluentTranslator)
    app.installTranslator(settingTranslator)
    myWindow = MyWindow()
    myWindow.show()
    app.exec_()
